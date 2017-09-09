import random
import math

import os


def thow_needle(needle_length, tile_seperation_distance):
    # I'm allowed to have the center placed between 0 and (.5 * tile_seperation_distance)
    # simply because I will have an "infinite" series of tiles to land on,
    # as such the model given in class will always apply.
    center_x = random.uniform(0, 0.5 * tile_seperation_distance)

    # I'm allowed to stop my random numbers at pi/2 because this would also fall in to the model shown in class.
    # There is no value added by allowing a full 180 degree or pi rotation simply because I could view the model
    # "upside down" in cases where there would be more than 90 degree or pi/2 rotation,
    # and as such I would effectively get the same output.
    theta = random.uniform(0, math.pi / 2)

    # Using cosine here to understand if the needle and the line are touching/crossing.
    tip_x = center_x - (0.5 * needle_length) * math.cos(theta)

    return bool(tip_x < 0)


def throw_needles(number):
    # I know I'm going to run 1000 throws but I want to be sure to get them split up nicely for Excel output
    throws = 0
    hits = 0
    data=[("Iteration","Hits","Throws","Estimate of Pi")]
    iterations=int((number / 50))

    for iteration in range(int((number / 50))):
        for i in range(50):
            needle_length = 10
            tile_seperation_distance = 10
            if (thow_needle(needle_length, tile_seperation_distance)):
                hits += 1
            throws += 1
        if (hits==0):
            # pi = 0
            # print("Iteration: " + str(iteration+1) + " Hits: " + str(hits) + " Throws: " + str(throws) + " Pi: " + str(pi))
            data.append((str(iteration+1), str(hits), str(throws), "0"))
            # Excel write function here
        else:
            pi = 2 * (throws / hits)
            # print("Iteration: " + str(iteration+1) + " Hits: " + str(hits) + " Throws: " + str(throws) + " Pi: " + str(pi))
            data.append((str(iteration + 1), str(hits), str(throws), str(pi)))
            # Excel write function here
    return data



def print_data_to_excel(data):
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, Series, Reference

    def create_chart(data):
        workbook = Workbook()
        ws = workbook.get_active_sheet()

        for row in data:
            if data.index(row)==0:
                iteration, hits, throws, pi = row
                ws.append((iteration, pi, "Actual Value of Pi"))
                continue

            iteration, hits, throws, pi = row
            addPi=float(pi)
            ws.append((iteration, addPi, math.pi))

        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Buffon's Needle"
        chart1.y_axis.title = 'Estimate of Pi'
        chart1.x_axis.title = 'Iteration'

        reference_data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=ws.max_row)
        categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

        chart1.add_data(reference_data, titles_from_data=True)
        chart1.set_categories(categories)

        chart1.shape = 4
        ws.add_chart(chart1, "F1")

        workbook.save("piEstimate.xlsx")
        workbook.close()



    current_files=os.listdir()

    if ("piEstimate.xlsx" in current_files):
        os.remove("piEstimate.xlsx")
        create_chart(data)
    else:
        create_chart(data)


def print_data(data):
    tuple_for_excel = data.pop(0)
    for row in data:
        iteration, hits, throws, pi=row
        print("Iteration: " + iteration + " Hits: " + hits + " Throws: " + throws + " Pi: " + pi)
    data.insert(0, tuple_for_excel)

if __name__=="__main__":
    data=throw_needles(1000)
    print_data(data)
    # try:
    print_data_to_excel(data)
    # except Exception:
    #     print("Something went wrong with the push to Excel. Please be sure that you have openpyxl installed as a "
    #           "python package.\n"
    #           "This can be achieved by running \"pip install openpyxl\" on your machine. \n\n"
    #           "If you have the piEstimate Excel document open, then simply close it and re-run.")